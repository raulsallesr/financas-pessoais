"""Importação segura e em memória da planilha de posição da B3."""

from __future__ import annotations

import re
import unicodedata
import warnings
from dataclasses import dataclass
from io import BytesIO
from math import isfinite
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_ARQUIVO_BYTES = 5_000_000

_ABAS_SUPORTADAS = {
    "Acoes",
    "Empréstimos",
    "ETF",
    "Fundo de Investimento",
    "Renda Fixa",
    "Tesouro Direto",
}

_COLUNAS_VALOR = {
    "Acoes": ("Valor Atualizado",),
    "Empréstimos": ("Valor Atualizado",),
    "ETF": ("Valor Atualizado",),
    "Fundo de Investimento": ("Valor Atualizado",),
    "Renda Fixa": (
        "Valor Atualizado MTM",
        "Valor Atualizado CURVA",
        "Valor Atualizado FECHAMENTO",
    ),
    "Tesouro Direto": (
        "Valor Atualizado",
        "Valor líquido",
        "Valor bruto",
    ),
}
_CAMPOS_NECESSARIOS = {
    "Produto",
    "Código de Negociação",
    "Código",
    "Indexador",
    *(
        coluna
        for colunas in _COLUNAS_VALOR.values()
        for coluna in colunas
    ),
}

_ETFS_CRIPTO = {"COIN11", "QBTC11", "HASH11"}
_ETFS_OURO = {"GOLD11"}
_ETFS_FII = {"XFIX11"}
_ETFS_RENDA_FIXA = {"NLFA11"}
_ETFS_EXTERIOR = {
    "ACWI11",
    "ALUG11",
    "IVVB11",
    "NASD11",
    "QQQI11",
    "SPYI11",
    "WRLD11",
}


class ErroImportacaoB3(Exception):
    """A planilha não pôde ser validada ou interpretada com segurança."""


@dataclass(frozen=True)
class PosicaoImportada:
    ativo: str
    classe: str
    valor_atual: float
    benchmark: str

    def para_editor(self) -> dict[str, object]:
        return {
            "Ativo": self.ativo,
            "Classe": self.classe,
            "Valor atual (R$)": self.valor_atual,
            "Valor investido (R$)": None,
            "Comparar com": self.benchmark,
        }


@dataclass(frozen=True)
class ResultadoImportacaoB3:
    posicoes: tuple[PosicaoImportada, ...]
    linhas_validas: int
    linhas_ignoradas: int
    abas_lidas: tuple[str, ...]

    @property
    def valor_total(self) -> float:
        return sum(posicao.valor_atual for posicao in self.posicoes)


def _texto(valor: object) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _numero(valor: object) -> float | None:
    if valor is None or (
        isinstance(valor, str) and valor.strip() in {"", "-"}
    ):
        return None
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return None
    return numero if isfinite(numero) else None


def _codigo_curto(*valores: object) -> str:
    textos = [_texto(valor) for valor in valores if _texto(valor)]
    if not textos:
        return ""
    texto = textos[0]
    prefixo = texto.split(" - ", 1)[0].strip()
    if re.fullmatch(r"[A-Z0-9]{3,12}", prefixo):
        return prefixo
    return texto[:120]


def _normalizar(texto: str) -> str:
    return "".join(
        caractere
        for caractere in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(caractere)
    ).upper()


def _classificar(
    aba: str,
    ativo: str,
    produto: str,
    indexador: str,
) -> tuple[str, str]:
    texto = _normalizar(f"{ativo} {produto} {indexador}")

    if "FUNDO DE INVESTIMENTO IMOBILIARIO" in texto or aba == (
        "Fundo de Investimento"
    ):
        return "Fundos imobiliários / FIAGRO", "Sem comparação"

    if aba in {"Acoes", "Empréstimos"}:
        return "Bolsa brasileira", "Sem comparação"

    if aba == "ETF":
        if ativo in _ETFS_CRIPTO or any(
            termo in texto for termo in ("BITCOIN", "CRYPTO", "CRIPTO")
        ):
            return "Bitcoin / cripto", "Bitcoin"
        if ativo in _ETFS_OURO or "OURO" in texto or "GOLD" in texto:
            return "Commodities / energia", "Sem comparação"
        if ativo in _ETFS_FII or "IFIX" in texto:
            return "Fundos imobiliários / FIAGRO", "Sem comparação"
        if ativo in _ETFS_RENDA_FIXA:
            return "Renda fixa pós-fixada", "CDI"
        if ativo in _ETFS_EXTERIOR:
            return "Exterior / dólar", "Dólar PTAX"
        return "Bolsa brasileira", "Sem comparação"

    if aba == "Tesouro Direto":
        if any(termo in texto for termo in ("IPCA", "RENDA+")):
            return "Títulos IPCA+", "Sem comparação"
        if "SELIC" in texto:
            return "Renda fixa pós-fixada", "Selic"
        return "Renda fixa prefixada", "Sem comparação"

    if aba == "Renda Fixa":
        if any(termo in texto for termo in ("CDI", "SELIC")):
            return "Renda fixa pós-fixada", "CDI"
        if any(termo in texto for termo in ("IPCA", "IGP-M", "IGPM")):
            return "Títulos IPCA+", "Sem comparação"
        return "Renda fixa prefixada", "Sem comparação"

    return "Outros", "Sem comparação"


def _linha_para_posicao(
    aba: str,
    registro: dict[str, object],
) -> PosicaoImportada | None:
    produto = _texto(registro.get("Produto"))
    codigo = _codigo_curto(
        registro.get("Código de Negociação"),
        registro.get("Código"),
        produto,
    )
    if not produto and not codigo:
        return None

    valor = next(
        (
            numero
            for coluna in _COLUNAS_VALOR[aba]
            if (numero := _numero(registro.get(coluna))) is not None
        ),
        None,
    )
    if valor is None or valor <= 0:
        return None

    ativo = codigo or produto[:120]
    classe, benchmark = _classificar(
        aba,
        ativo,
        produto,
        _texto(registro.get("Indexador")),
    )
    return PosicaoImportada(
        ativo=ativo,
        classe=classe,
        valor_atual=round(valor, 2),
        benchmark=benchmark,
    )


def importar_posicao_b3(conteudo: bytes) -> ResultadoImportacaoB3:
    """Lê somente campos financeiros necessários; não persiste o arquivo."""
    if not conteudo:
        raise ErroImportacaoB3("A planilha enviada está vazia.")
    if len(conteudo) > MAX_ARQUIVO_BYTES:
        raise ErroImportacaoB3(
            "A planilha excede o limite de 5 MB para importação local."
        )

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style.*",
                category=UserWarning,
                module="openpyxl.styles.stylesheet",
            )
            workbook = load_workbook(
                BytesIO(conteudo),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
    except (
        BadZipFile,
        EOFError,
        InvalidFileException,
        OSError,
        ValueError,
    ) as erro:
        raise ErroImportacaoB3(
            "O arquivo não é uma planilha XLSX válida."
        ) from erro

    agregadas: dict[tuple[str, str, str], float] = {}
    linhas_validas = 0
    linhas_ignoradas = 0
    abas_lidas: list[str] = []
    try:
        for aba in workbook.sheetnames:
            if aba not in _ABAS_SUPORTADAS:
                continue
            planilha = workbook[aba]
            if planilha.max_row == 1 and planilha.max_column == 1:
                # A exportação atual da B3 declara dimensão A1 mesmo quando
                # há dezenas de linhas. O modo read_only confia nesse valor.
                planilha.reset_dimensions()
            linhas = planilha.iter_rows(values_only=True)
            cabecalho = next(linhas, None)
            if not cabecalho or "Produto" not in cabecalho:
                continue
            colunas = [
                _texto(valor) if valor is not None else ""
                for valor in cabecalho
            ]
            indices_necessarios = {
                indice: coluna
                for indice, coluna in enumerate(colunas)
                if coluna in _CAMPOS_NECESSARIOS
            }
            abas_lidas.append(aba)
            for valores in linhas:
                registro = {
                    coluna: valores[indice]
                    for indice, coluna in indices_necessarios.items()
                    if indice < len(valores)
                }
                posicao = _linha_para_posicao(aba, registro)
                if posicao is None:
                    linhas_ignoradas += 1
                    continue
                linhas_validas += 1
                chave = (
                    posicao.ativo,
                    posicao.classe,
                    posicao.benchmark,
                )
                agregadas[chave] = (
                    agregadas.get(chave, 0.0) + posicao.valor_atual
                )
    finally:
        workbook.close()

    if not abas_lidas:
        raise ErroImportacaoB3(
            "Nenhuma aba reconhecida da posição B3 foi encontrada."
        )
    if not agregadas:
        raise ErroImportacaoB3(
            "A planilha não contém posições com valor atualizado positivo."
        )

    posicoes = tuple(
        PosicaoImportada(
            ativo=ativo,
            classe=classe,
            valor_atual=round(valor, 2),
            benchmark=benchmark,
        )
        for (ativo, classe, benchmark), valor in sorted(
            agregadas.items(),
            key=lambda item: (-item[1], item[0][0]),
        )
    )
    return ResultadoImportacaoB3(
        posicoes=posicoes,
        linhas_validas=linhas_validas,
        linhas_ignoradas=linhas_ignoradas,
        abas_lidas=tuple(abas_lidas),
    )
