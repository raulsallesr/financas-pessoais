import sys
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from focuslens.adapters.b3_importacao import ErroImportacaoB3, importar_posicao_b3


def _planilha_sintetica() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    acoes = workbook.create_sheet("Acoes")
    acoes.append(
        ["Produto", "Código de Negociação", "Valor Atualizado"]
    )
    acoes.append(["Empresa energia", "EGIE3 - EMPRESA", 1_000])
    acoes.append(["Mineradora", "VALE3 - EMPRESA", 200])
    acoes.append([None, None, 1_200])  # subtotal da exportação

    emprestimos = workbook.create_sheet("Empréstimos")
    emprestimos.append(["Produto", "Valor Atualizado"])
    emprestimos.append(["VALE3 - EMPRESA", 300])
    emprestimos.append([None, 300])  # subtotal da exportação

    etf = workbook.create_sheet("ETF")
    etf.append(
        ["Produto", "Código de Negociação", "Valor Atualizado"]
    )
    etf.append(["ETF exterior", "IVVB11 - FUNDO", 400])
    etf.append(["ETF cripto", "HASH11 - FUNDO", 500])
    etf.append(["ETF ouro", "GOLD11 - FUNDO", 600])
    etf.append(["ETF renda fixa", "NLFA11 - FUNDO", 700])

    fundos = workbook.create_sheet("Fundo de Investimento")
    fundos.append(
        ["Produto", "Código de Negociação", "Valor Atualizado"]
    )
    fundos.append(["FII", "MXRF11 - FUNDO", 800])

    renda_fixa = workbook.create_sheet("Renda Fixa")
    renda_fixa.append(
        [
            "Produto",
            "Código",
            "Indexador",
            "Valor Atualizado MTM",
            "Valor Atualizado CURVA",
        ]
    )
    renda_fixa.append(
        [
            "CFF - FUNDO DE INVESTIMENTO IMOBILIARIO",
            None,
            "-",
            "-",
            900,
        ]
    )
    renda_fixa.append(["CDB", "CDB123", "CDI", 1_000, 1_010])

    tesouro = workbook.create_sheet("Tesouro Direto")
    tesouro.append(["Produto", "Valor Atualizado"])
    tesouro.append(["Tesouro IPCA+ 2035", 1_100])
    tesouro.append(["Tesouro Selic 2029", 1_200])

    saida = BytesIO()
    workbook.save(saida)
    workbook.close()
    return saida.getvalue()


def _forcar_dimensao_a1(conteudo: bytes) -> bytes:
    origem = BytesIO(conteudo)
    destino = BytesIO()
    with (
        ZipFile(origem, "r") as entrada,
        ZipFile(destino, "w", ZIP_DEFLATED) as saida,
    ):
        for item in entrada.infolist():
            dados = entrada.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                dados = re.sub(
                    rb'<dimension ref="[^"]+"',
                    b'<dimension ref="A1"',
                    dados,
                    count=1,
                )
            saida.writestr(item, dados)
    return destino.getvalue()


def test_importa_consolida_e_ignora_subtotais_da_b3():
    resultado = importar_posicao_b3(_planilha_sintetica())

    assert len(resultado.posicoes) == 11
    assert resultado.linhas_validas == 12
    assert resultado.linhas_ignoradas == 2
    assert resultado.valor_total == 8_700

    por_ativo = {posicao.ativo: posicao for posicao in resultado.posicoes}
    assert por_ativo["VALE3"].valor_atual == 500
    assert por_ativo["IVVB11"].classe == "Exterior / dólar"
    assert por_ativo["HASH11"].classe == "Bitcoin / cripto"
    assert por_ativo["GOLD11"].classe == "Commodities / energia"
    assert por_ativo["NLFA11"].classe == "Renda fixa pós-fixada"
    assert por_ativo["MXRF11"].classe == (
        "Fundos imobiliários / FIAGRO"
    )
    assert por_ativo["CFF"].valor_atual == 900
    assert por_ativo["CFF"].classe == "Fundos imobiliários / FIAGRO"
    assert por_ativo["Tesouro IPCA+ 2035"].classe == "Títulos IPCA+"
    assert por_ativo["Tesouro Selic 2029"].benchmark == "Selic"


def test_importa_exportacao_b3_com_dimensao_a1_incorreta():
    conteudo = _forcar_dimensao_a1(_planilha_sintetica())
    resultado = importar_posicao_b3(conteudo)

    assert len(resultado.posicoes) == 11
    assert resultado.valor_total == 8_700


def test_importacao_nao_le_identificadores_que_nao_sao_necessarios():
    resultado = importar_posicao_b3(_planilha_sintetica())
    registros = [posicao.para_editor() for posicao in resultado.posicoes]

    assert all(
        set(registro)
        == {
            "Ativo",
            "Classe",
            "Valor atual (R$)",
            "Valor investido (R$)",
            "Comparar com",
        }
        for registro in registros
    )


@pytest.mark.parametrize("conteudo", [b"", b"nao e um xlsx"])
def test_rejeita_arquivo_vazio_ou_invalido(conteudo):
    with pytest.raises(ErroImportacaoB3):
        importar_posicao_b3(conteudo)


def test_rejeita_planilha_sem_abas_da_posicao_b3():
    workbook = Workbook()
    workbook.active.title = "Outra aba"
    saida = BytesIO()
    workbook.save(saida)
    workbook.close()

    with pytest.raises(ErroImportacaoB3, match="Nenhuma aba reconhecida"):
        importar_posicao_b3(saida.getvalue())
